# coding=utf-8
import openpyxl
import os

base_path = os.path.dirname(os.getcwd())
excel_path = base_path + "/Case/cwz.xlsx"


class HandleExcel:
    # 加载Excel文件
    def load_excel(self):
        open_excel = openpyxl.load_workbook(excel_path)
        return open_excel

    # 获取sheet
    def get_sheet_data(self, index=None):
        open_excel = self.load_excel()
        sheet_name = open_excel.sheetnames
        if index is None:
            index = 0
        excel_value = open_excel[sheet_name[index]]
        return excel_value

    # 获取单元格内容
    def get_cell_value(self, row, col):
        excel_value = self.get_sheet_data()
        cell_value = excel_value.cell(row=row, column=col).value
        return cell_value

    # 获取最大行数
    def get_rows(self):
        excel_value = self.get_sheet_data()
        rows = excel_value.max_row
        return rows

    # 获取行的内容
    def get_rows_value(self, row):
        row_list = []
        excel_value = self.get_sheet_data()[row]
        for i in excel_value:
            row_list.append(i.value)
        return row_list


if __name__ == "__main__":
    handle_excel = HandleExcel()
    print(handle_excel.get_cell_value(1, 1))
    print(handle_excel.get_rows())
